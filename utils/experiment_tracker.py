"""
Experiment tracker: saves JSON results and exports multi-sheet Excel reports.

Excel sheets:
  1. Overall Raw Metrics        — per model
  2. Aggregated Scores          — per model × 4 strategies
  3. Per-Dataset Breakdown      — per model × per dataset
  4. Experiment Metadata
"""

import json
import os
import time
from typing import Any, Dict, List

from utils.logger import get_logger

logger = get_logger(__name__)

METRIC_KEYS   = ["correctness", "consistency", "robustness",
                 "logical_coherence", "efficiency", "stability"]
# Fallback order for known strategies; actual columns are derived dynamically
# from the results so ALL strategies in the config appear in the Excel output.
STRATEGY_KEYS = ["balanced", "safety_priority", "accuracy_priority", "efficiency_priority"]


def _strategy_keys(results):
    """All aggregation strategy keys present in results, stable order:
    known keys first (STRATEGY_KEYS order), then any extras in first-seen order."""
    seen = []
    for r in results:
        for k in (r.get("aggregated") or {}):
            if k not in seen:
                seen.append(k)
    known = [k for k in STRATEGY_KEYS if k in seen]
    extra = [k for k in seen if k not in STRATEGY_KEYS]
    return known + extra
METRIC_LABELS = ["Correctness", "Consistency", "Robustness",
                 "Logical Coherence", "Efficiency", "Stability"]


class ExperimentTracker:

    def __init__(self, experiment_id: str, output_dir: str = "outputs"):
        self.experiment_id = experiment_id
        self.output_dir    = output_dir
        self.exp_dir       = os.path.join(output_dir, experiment_id)
        os.makedirs(self.exp_dir, exist_ok=True)
        self._model_results: List[Dict[str, Any]] = []
        self._start_time    = time.time()

    def log_model_result(self, model_name: str, result: Dict[str, Any]) -> None:
        self._model_results.append(result)
        os.makedirs(self.exp_dir, exist_ok=True)  # ensure dir exists
        safe_name = model_name.replace("/", "_").replace("\\", "_")
        path = os.path.join(self.exp_dir, f"{safe_name}_result.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        logger.info(f"Saved: {path}")

    def save_summary(self, results: List[Dict[str, Any]]) -> str:
        summary = {
            "experiment_id":    self.experiment_id,
            "timestamp":        time.strftime("%Y-%m-%dT%H:%M:%S"),
            "duration_seconds": round(time.time() - self._start_time, 2),
            "num_models":       len(results),
            "results":          results,
        }
        path = os.path.join(self.exp_dir, "summary.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        logger.info(f"Summary saved: {path}")
        return path

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_excel(
        self,
        results: List[Dict[str, Any]],
        filename: str = "results.xlsx",
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not available. Skipping Excel export.")
            return ""

        wb = openpyxl.Workbook()

        # Style helpers
        HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
        ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
        SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
        SUB_FONT  = Font(bold=True, color="FFFFFF", size=10)
        CTR       = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="AAAAAA")
        BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CTR; c.border = BORDER
            return c

        def cell(ws, row, col, value, alt=False, fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            if alt: c.fill = ALT_FILL
            c.alignment = CTR; c.border = BORDER
            if fmt: c.number_format = fmt
            return c

        # ---- Sheet 1: Overall Raw Metrics ----------------------------
        ws1 = wb.active
        ws1.title = "Overall Raw Metrics"
        headers = ["Model"] + METRIC_LABELS
        for col, h in enumerate(headers, 1):
            hdr(ws1, 1, col, h)
            ws1.column_dimensions[get_column_letter(col)].width = 18

        for ri, r in enumerate(results, 2):
            alt = ri % 2 == 0
            cell(ws1, ri, 1, r["model"], alt)
            for ci, key in enumerate(METRIC_KEYS, 2):
                v = r["raw_metrics"].get(key, "N/A")
                cell(ws1, ri, ci, v, alt, fmt="0.0000" if isinstance(v, float) else None)

        ws1.row_dimensions[1].height = 30
        ws1.freeze_panes = "B2"

        # ---- Sheet 2: Aggregated Scores ------------------------------
        ws2 = wb.create_sheet("Aggregated Scores")
        strat_keys = _strategy_keys(results)
        agg_headers = ["Model"] + [s.replace("_", " ").title() for s in strat_keys]
        for col, h in enumerate(agg_headers, 1):
            hdr(ws2, 1, col, h)
            ws2.column_dimensions[get_column_letter(col)].width = 22

        for ri, r in enumerate(results, 2):
            alt = ri % 2 == 0
            cell(ws2, ri, 1, r["model"], alt)
            for ci, key in enumerate(strat_keys, 2):
                v = r.get("aggregated", {}).get(key, "N/A")
                cell(ws2, ri, ci, v, alt, fmt="0.0000" if isinstance(v, float) else None)

        ws2.row_dimensions[1].height = 30
        ws2.freeze_panes = "B2"

        # ---- Sheet 3: Per-Dataset Breakdown --------------------------
        ws3 = wb.create_sheet("Per-Dataset Breakdown")
        pd_headers = ["Model", "Dataset", "# Samples"] + METRIC_LABELS + ["Balanced"]
        for col, h in enumerate(pd_headers, 1):
            hdr(ws3, 1, col, h)
            ws3.column_dimensions[get_column_letter(col)].width = 16

        ws3.column_dimensions["A"].width = 24
        ws3.column_dimensions["B"].width = 18

        row_num = 2
        for r in results:
            per_ds = r.get("per_dataset", {})
            if not per_ds:
                # No breakdown — write overall as single row
                alt = row_num % 2 == 0
                cell(ws3, row_num, 1, r["model"], alt)
                cell(ws3, row_num, 2, "overall", alt)
                cell(ws3, row_num, 3, r["metadata"].get("num_samples", ""), alt)
                for ci, key in enumerate(METRIC_KEYS, 4):
                    v = r["raw_metrics"].get(key, "N/A")
                    cell(ws3, row_num, ci, v, alt, fmt="0.0000" if isinstance(v, float) else None)
                v_bal = r.get("aggregated", {}).get("balanced", "N/A")
                cell(ws3, row_num, 4 + len(METRIC_KEYS), v_bal, alt,
                     fmt="0.0000" if isinstance(v_bal, float) else None)
                row_num += 1
            else:
                for ds_name in sorted(per_ds.keys()):
                    ds_data = per_ds[ds_name]
                    alt = row_num % 2 == 0
                    cell(ws3, row_num, 1, r["model"], alt)
                    cell(ws3, row_num, 2, ds_name, alt)
                    cell(ws3, row_num, 3, ds_data.get("num_samples", ""), alt)
                    for ci, key in enumerate(METRIC_KEYS, 4):
                        v = ds_data["raw_metrics"].get(key, "N/A")
                        cell(ws3, row_num, ci, v, alt,
                             fmt="0.0000" if isinstance(v, float) else None)
                    v_bal = ds_data.get("aggregated", {}).get("balanced", "N/A")
                    cell(ws3, row_num, 4 + len(METRIC_KEYS), v_bal, alt,
                         fmt="0.0000" if isinstance(v_bal, float) else None)
                    row_num += 1

        ws3.row_dimensions[1].height = 30
        ws3.freeze_panes = "C2"

        # ---- Sheet 4: Metadata ---------------------------------------
        ws4 = wb.create_sheet("Experiment Metadata")
        bold = Font(bold=True)
        meta_rows = [
            ("Experiment ID",     self.experiment_id),
            ("Timestamp",         time.strftime("%Y-%m-%dT%H:%M:%S")),
            ("Duration (s)",      round(time.time() - self._start_time, 2)),
            ("Models Evaluated",  len(results)),
        ]
        if results:
            meta_rows += [
                ("Total Samples",    results[0]["metadata"].get("num_samples", "N/A")),
                ("Datasets",         ", ".join(results[0]["metadata"].get("datasets", []))),
                ("Consistency Runs", results[0]["metadata"].get("consistency_runs", "N/A")),
                ("Robustness Perturb.", results[0]["metadata"].get("robustness_perturbations", "N/A")),
            ]

        for i, (k, v) in enumerate(meta_rows, 1):
            ws4.cell(row=i, column=1, value=k).font = bold
            ws4.cell(row=i, column=2, value=v)
        ws4.column_dimensions["A"].width = 30
        ws4.column_dimensions["B"].width = 40

        path = os.path.join(self.exp_dir, filename)
        wb.save(path)
        logger.info(f"Excel saved: {path}")
        return path